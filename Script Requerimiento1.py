# -*- coding: utf-8 -*-
"""
"""

#CABECERA-----------------------------------------------------------------------------
import requests
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from statsmodels.tsa.holtwinters import ExponentialSmoothing
from openpyxl import Workbook  # Agregar importación de Workbook
#DEFINICIONES/CONSTANTES-------------------------------------------------------------
# URLs de las series
urls =  {'Diaria': "https://www.banxico.org.mx/SieAPIRest/service/v1/series/SF63528,SP68257,SF60653/datos?token=fb1ce37630c5fbf74ff5bd6f2b1ea032eeb356419b4a45196c5027b5111708fb",
         'Mensual': "https://www.banxico.org.mx/SieAPIRest/service/v1/series/SF17801,SF221962,SF282,SF283,SF285,SF286,SF3270,SF3338,SF3366,SF3367,SF3368,SF43717,SF43785,SI561,SI562,SL11295,SL11298,SP1,SI563,SI564,SI264,SI222,SI223,SI260,SI224,SI261,SI220,SI263,SI209,SI262,SI225,SI237/datos?token=fb1ce37630c5fbf74ff5bd6f2b1ea032eeb356419b4a45196c5027b5111708fb"}

data = {'CCP':'SF286','CCP_USD':'SF3366','CCP_UDIS':'SF3368','Cetes182':'SF3270',
        'Cetes28':'SF282','Cetes364':'SF3367','Cetes91':'SF3338','CPP':'SF285',
        'Inflacion':'CP151','INPC':'SP1','Libor3M':'SI561','Libor6M':'SI562',
        'PIB':'CR200','SMG':'SL11298','SMZLDLFN':'SL11295','FONDEO_BANCARIO':'SF43773',
        'FX':'SF43717','TIIE182':'SF221962','TIIE28':'SF283','TIIE91':'SF17801',
        'UDIS':'SP68257','USD_MXN':'SF63528','USD_MXN_LIQ':'SF60653','lETRA_1M':'SI264','LETRA_3M':'SI222','LETRA_6M':'SI223','LETRA_1A':'SI260','LETRA_2A':'SI224',
        'LETRA_3A':'SI261','LETRA_5A':'SI220','LETRA_7A':'SI263','LETRA_10A':'SI209','LETRA_20A':'SI262','LETRA_30A':'SI225',}
        
        #'USD_MXN_LIQ':'SF60653','TASA_OBJETIVO':'SF61745',
        #'TIEE_FONDEO1':'SF331451','USD_MXN_V48':'SF43784',
        #'lETRAT_3M':'SI563','LETRAT_6M':'SI564',
        #
        #'TASA_FON_FED':'SI237'}
        #SF60653,SF61745,SF331451,SF43773,SF43784
        #

# Agrupar las series que quieres fusionar
TIIE = ['TIIE1', 'TIIE28', 'TIIE91','TIIE1821']
CETES = ['T_CETES28', 'T_CETES91', 'T_CETES182']

# Crear un nuevo libro de trabajo de Excel
wb = Workbook()

#FUNCIONES---------------------------------------------------------------------------
def to_month(url, periodo): #definir el nombre de la función con parámetros de entrada 
    response = requests.get(url)
    if response.status_code == 200:#validar comunicación web correcta
        try:
            data = response.json()
        except ValueError as e:
            print(f"Error al decodificar JSON ({periodo}): {e}")
            return None
        
        series = data.get("bmx", {}).get("series", [])
        if not series:
            print(f"Error: No se encontraron series en los datos ({periodo})")
            return None

        series_dict = {}
        for serie_data in series:
            series_list = serie_data.get('datos', [])
            if not series_list:
                print(f"Error: No se encontraron datos en la serie ({periodo})")
                continue
            
            df_id = str(serie_data.get('idSerie', 'Unknown'))#identificador de serie
            df_name = ApiID_to_name.get(df_id)
            df = pd.DataFrame(series_list)
            df['dato'] = pd.to_numeric(df['dato'], errors='coerce')
            df['fecha'] = pd.to_datetime(df['fecha'], errors='coerce', dayfirst=True)
            df = df.dropna(subset=['dato'])

            if df.empty:
                print(f"Error: No se pudo procesar la serie '{df_name}' ({periodo})")
                continue

            if periodo in ('Diaria', 'Semanal'):
                df['fecha'] = df['fecha'].dt.to_period('M').dt.to_timestamp()
                df = df.groupby('fecha')['dato'].mean().reset_index()
            series_dict[df_name] = pd.Series(df['dato'].values, index=df['fecha'])

        return series_dict
    else:
        print(f"Error en la solicitud ({periodo}): {response.status_code}")
        return None
    #BODY--------------------------------------------------------------------------------

data = {'Params': data.keys(), 'ApiID': data.values()} # Crear un diccionario con tus datos
df_names = pd.DataFrame(data) # Crear un DataFrame a partir del diccionario
# Crea un diccionario que asocie el ID de la API con el nombre del DataFrame
ApiID_to_name = dict(zip(df_names['ApiID'], df_names['Params']))

#------------------------------------------------------------------------

#--------------------------------------------------------------------------

# Obtener y procesar los datos para cada periodo: Daily, Weekly y monthly
for periodo, url in urls.items():
    series = to_month(url, periodo)  # Agregué esta línea para definir la variable 'series'
    print(series)
    for serie_name, serie in series.items():
        # Filtrar la serie a solo 5 años
        serie = serie[serie.index >= serie.index[-1] - pd.DateOffset(years=5)]#Modificar a 6 años
        # Dividir la serie temporal en entrenamiento y prueba
        train_size = int(len(serie) * (1))
        train = serie[:train_size]
        
        # Aplicar el modelo de Holt-Winters
        modelo = ExponentialSmoothing(train, trend='add', seasonal='add', seasonal_periods=12)
        resultado_modelo = modelo.fit()
        
        # Realizar predicciones
        predicciones = resultado_modelo.forecast(6)
        
        # Crear una lista de tuplas para los datos de entrenamiento y predicciones
        train_data = [(index.strftime('%Y-%m-%d'), value) for index, value in train.items() if isinstance(index, pd.Timestamp)]
        predicciones_data = [(index.strftime('%Y-%m-%d'), value) for index, value in predicciones.items() if isinstance(index, pd.Timestamp)]
        
        data_rows = train_data + predicciones_data
        
        # Graficar resultados de Entrenamiento y Predicciones juntos
        plt.figure(figsize=(12, 6))
        plt.plot(train.index, train.values, label='Entrenamiento')
        plt.plot(predicciones.index, predicciones.values, label='Predicciones', linestyle='dashed', color='red')
        plt.title(f'Datos de Entrenamiento y Predicciones - {serie_name}')
        plt.legend()
        plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
        plt.gca().xaxis.set_major_locator(mdates.MonthLocator(interval=2))  # Agregar más fechas de referencia
        plt.gca().grid(True)  # Agregar una malla
        plt.xticks(rotation=50, ha='right')  # Rotar las fechas n grados en el eje X
        plt.show()
        
        # Escribir los datos de Entrenamiento y Predicciones en el archivo de Excel
        ws = wb.create_sheet(serie_name)
        ws.append(['Fecha', serie_name])  # Agregar el encabezado directamente aquí
        for row in data_rows:
            ws.append(row)
# Eliminar la hoja por defecto
default_sheet = wb['Sheet']
wb.remove(default_sheet)

# Guardar el libro de trabajo en un archivo Excel
wb.save('all_series.xlsx')


# Unir las series de TIIE y CETES
tiie_series = ['TIIE28', 'TIIE91', 'TIIE182']
cetes_series = ['Cetes28', 'Cetes91', 'Cetes364']
series_to_merge = tiie_series + cetes_series

merged_series = {}
for serie_name in series_to_merge:
    if serie_name in tiie_series:
        merged_series['TIIE'] = merged_series.get('TIIE', pd.Series()).add(series[serie_name], fill_value=0)
    elif serie_name in cetes_series:
        merged_series['CETES'] = merged_series.get('CETES', pd.Series()).add(series[serie_name], fill_value=0)

# Graficar las series fusionadas
for serie_name, serie in merged_series.items():
    plt.figure(figsize=(12, 6))
    plt.plot(serie.index, serie.values, label=serie_name)
    plt.title(f'Datos Fusionados - {serie_name}')
    plt.legend()
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
    plt.gca().xaxis.set_major_locator(mdates.MonthLocator(interval=1))
    plt.gca().grid(True)
    plt.xticks(rotation=50, ha='right')
    plt.show()

    # Escribir los datos fusionados en el archivo de Excel
    ws = wb.create_sheet(serie_name)
    ws.append(['Fecha', serie_name])  # Agregar el encabezado directamente aquí
    for index, value in serie.items():
        ws.append([index.strftime('%Y-%m-%d'), value])

# Eliminar la hoja por defecto
#default_sheet = wb['Sheet']
#wb.remove(default_sheet)

# Guardar el libro de trabajo en un archivo Excel
wb.save('merged_series.xlsx')

##############PRINT##############



# Leer el archivo Excel
ruta_archivo = 'merged_series.xlsx'
df = pd.read_excel(ruta_archivo)

# Mostrar el DataFrame
print(df)